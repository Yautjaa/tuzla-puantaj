"""
Excel işleme modülü — UploadThing URL'den izin.xlsx indirir ve işler.
Vercel serverless (read-only filesystem) uyumludur.
"""
import unicodedata
from io import BytesIO

import httpx
import openpyxl


def normalize(s: str) -> str:
    s = unicodedata.normalize("NFC", s)
    s = s.replace("İ", "I").replace("ı", "i")
    return s.upper().strip()


def read_izin_from_url(file_url: str) -> list[dict]:
    """UploadThing URL'den izin.xlsx indir ve personelleri oku."""
    response = httpx.get(file_url, timeout=30.0)
    response.raise_for_status()

    wb = openpyxl.load_workbook(BytesIO(response.content))
    ws = wb.active
    personeller = []

    for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if len(row) < 2:
            continue
        sira = row[0] if row[0] is not None else idx
        ad_soyad = row[1]
        izin_gunu = row[2] if len(row) > 2 else ""

        if not ad_soyad:
            continue

        ad_soyad_str = str(ad_soyad).strip()
        if ad_soyad_str.lower() in [
            "ad soyad", "ad", "soyad", "isim", "ad-soyad",
            "adi soyadi", "adı soyadı", "personel", "personel adı",
        ]:
            continue

        izin_gun_str = str(izin_gunu) if izin_gunu is not None else ""

        try:
            personel_id = int(sira)
        except ValueError:
            personel_id = idx

        personeller.append({
            "id": personel_id,
            "ad_soyad": normalize(ad_soyad_str),
            "izin_gunu": normalize(izin_gun_str),
        })

    return personeller


def read_izin_from_bytes(file_bytes: bytes) -> list[dict]:
    """Doğrudan byte verisinden izin.xlsx oku (upload sırasında kullanılır)."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws = wb.active
    personeller = []

    for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if len(row) < 2:
            continue
        sira = row[0] if row[0] is not None else idx
        ad_soyad = row[1]
        izin_gunu = row[2] if len(row) > 2 else ""

        if not ad_soyad:
            continue

        ad_soyad_str = str(ad_soyad).strip()
        if ad_soyad_str.lower() in [
            "ad soyad", "ad", "soyad", "isim", "ad-soyad",
            "adi soyadi", "adı soyadı", "personel", "personel adı",
        ]:
            continue

        izin_gun_str = str(izin_gunu) if izin_gunu is not None else ""

        try:
            personel_id = int(sira)
        except ValueError:
            personel_id = idx

        personeller.append({
            "id": personel_id,
            "ad_soyad": normalize(ad_soyad_str),
            "izin_gunu": normalize(izin_gun_str),
        })

    return personeller


def puantaj_excel_olustur(yil: int, ay: int, puantaj_data: list[dict], template_bytes: bytes) -> bytes:
    """
    Template'den puantaj Excel dosyası oluştur.
    template_bytes: gömülü template dosyasının byte verisi.
    """
    import calendar
    import datetime

    wb = openpyxl.load_workbook(BytesIO(template_bytes))
    ws = wb["Puantaj Bilgileri"]

    _, aydaki_gun_sayisi = calendar.monthrange(yil, ay)

    # Başlık satırındaki tarih sütunlarını dinamik oluştur
    tarih_sutunlari: dict[int, int] = {}
    for gun_no in range(1, 32):
        col_idx = gun_no + 3
        cell = ws.cell(row=1, column=col_idx)

        if gun_no <= aydaki_gun_sayisi:
            cell.value = datetime.datetime(yil, ay, gun_no)
            tarih_sutunlari[gun_no] = col_idx
        else:
            cell.value = ""

    # Her satırdaki personeli bul ve güncelle
    for row_idx in range(2, ws.max_row + 1):
        ad = ws.cell(row=row_idx, column=1).value
        soyad = ws.cell(row=row_idx, column=2).value
        if not ad:
            break

        tam_isim = normalize(f"{ad} {soyad}")

        personel_data = next(
            (p for p in puantaj_data if normalize(p["ad_soyad"]) == tam_isim),
            None,
        )

        if personel_data:
            for gun_no_str, deger in personel_data["gunler"].items():
                col = tarih_sutunlari.get(int(gun_no_str))
                if col:
                    ws.cell(row=row_idx, column=col).value = deger

            for gun_no in range(aydaki_gun_sayisi + 1, 32):
                col_idx = gun_no + 3
                ws.cell(row=row_idx, column=col_idx).value = ""

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
