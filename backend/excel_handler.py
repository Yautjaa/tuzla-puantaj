import os
import unicodedata
from io import BytesIO

import openpyxl

IZIN_PATH = os.path.join(os.path.dirname(__file__), "data", "izin.xlsx")
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "templates", "liste_template.xlsx")


def normalize(s: str) -> str:
    """Türkçe karakter sorunlarına karşı normalize et; İ→I, ı→i, sonra büyüt."""
    s = unicodedata.normalize("NFC", s)
    s = s.replace("İ", "I").replace("ı", "i")
    return s.upper().strip()


def read_izin_xlsx() -> list[dict]:
    """
    Returns:
        [{"id": 1, "ad_soyad": "AYŞE ARSLAN", "izin_gunu": "PERŞEMBE"}, ...]
    """
    if not os.path.exists(IZIN_PATH):
        return []

    wb = openpyxl.load_workbook(IZIN_PATH)
    ws = wb.active
    personeller = []

    for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if len(row) < 2:
            continue
        sira = row[0] if row[0] is not None else idx
        ad_soyad = row[1]
        izin_gunu = row[2] if len(row) > 2 else ""

        if not ad_soyad:
            continue
            
        ad_soyad_str = str(ad_soyad).strip()
        if ad_soyad_str.lower() in ["ad soyad", "ad", "soyad", "isim", "ad-soyad", "adi soyadi", "adı soyadı", "personel", "personel adı"]:
            continue

        izin_gun_str = str(izin_gunu) if izin_gunu is not None else ""

        try:
            personel_id = int(sira)
        except ValueError:
            personel_id = idx

        personeller.append({
            "id": personel_id,
            "ad_soyad": normalize(ad_soyad_str),
            "izin_gunu": normalize(izin_gun_str),
        })

    return personeller


def write_back_izin(personeller: list[dict]) -> None:
    """İzin günü değişikliklerini izin.xlsx'e geri yaz."""
    if not os.path.exists(IZIN_PATH):
        return

    wb = openpyxl.load_workbook(IZIN_PATH)
    ws = wb.active

    guncelleme_map = {p["id"]: p["izin_gunu"] for p in personeller}

    for idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        if len(row) < 2:
            continue
        sira_cell = row[0]
        ad_soyad_cell = row[1]
        
        # Ensure row has at least 3 columns for writing back izin_gunu
        if len(row) < 3:
            # Need to extend row or just add value, but iterating doesn't easily create cells if missing.
            # We can use ws.cell(row=row_idx, column=3) below.
            pass
            
        ad_soyad = ad_soyad_cell.value
        if not ad_soyad:
            continue
            
        ad_soyad_str = str(ad_soyad).strip()
        if ad_soyad_str.lower() in ["ad soyad", "ad", "soyad", "isim", "ad-soyad", "adi soyadi", "adı soyadı", "personel", "personel adı"]:
            continue
            
        sira = sira_cell.value if sira_cell.value is not None else idx
        try:
            personel_id = int(sira)
        except ValueError:
            personel_id = idx

        if personel_id in guncelleme_map:
            # Re-read the row index from the cell since idx might not match exactly if there are skipped rows
            ws.cell(row=ad_soyad_cell.row, column=3).value = guncelleme_map[personel_id]

    wb.save(IZIN_PATH)


def puantaj_excel_olustur(yil: int, ay: int, puantaj_data: list[dict]) -> bytes:
    """
    Template'i kopyala, puantaj verisiyle doldur, bytes olarak döndür.
    puantaj_data: [{"ad_soyad": "...", "gunler": {"1": "X", ...}}, ...]
    """
    import calendar
    import datetime
    
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Puantaj Bilgileri"]

    _, aydaki_gun_sayisi = calendar.monthrange(yil, ay)

    # Başlık satırındaki tarih sütunlarını dinamik oluştur
    tarih_sutunlari: dict[int, int] = {}
    for gun_no in range(1, 32):
        col_idx = gun_no + 3
        cell = ws.cell(row=1, column=col_idx)
        
        if gun_no <= aydaki_gun_sayisi:
            cell.value = datetime.datetime(yil, ay, gun_no)
            tarih_sutunlari[gun_no] = col_idx
        else:
            # Fazla günleri temizle (örn. Şubat için 29, 30, 31)
            cell.value = ""

    # Her satırdaki personeli bul ve güncelle
    for row_idx in range(2, ws.max_row + 1):
        ad = ws.cell(row=row_idx, column=1).value
        soyad = ws.cell(row=row_idx, column=2).value
        if not ad:
            break

        tam_isim = normalize(f"{ad} {soyad}")

        personel_data = next(
            (p for p in puantaj_data if normalize(p["ad_soyad"]) == tam_isim),
            None,
        )

        if personel_data:
            for gun_no_str, deger in personel_data["gunler"].items():
                col = tarih_sutunlari.get(int(gun_no_str))
                if col:
                    ws.cell(row=row_idx, column=col).value = deger
                    
            # Personelin fazla gün hücrelerini (X vs) temizle
            for gun_no in range(aydaki_gun_sayisi + 1, 32):
                col_idx = gun_no + 3
                ws.cell(row=row_idx, column=col_idx).value = ""

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
